import unittest
from unittest.mock import Mock, patch, MagicMock, mock_open
import sys, os
import json
from openpyxl import Workbook

# Add the directory containing excel_management.py to the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'scripts')))


from excel_management import (
    preProcess, insertData, insertError, create_excel,
    modelProcess, manuProcess, dateProcess, swlProcess, jsonProcess
)

class TestExcelManagement(unittest.TestCase):

    def setUp(self):
        # This runs before each test
        self.workbook = Workbook()
        self.sheet_data = self.workbook.active

    def test_preProcess(self):
        column_mapping = {
            "Id Number": "A",
            "RFID": "B",
            "Item Category": "C"
        }
        preProcess(self.sheet_data, column_mapping)
        for header, column in column_mapping.items():
            cell = self.sheet_data[column + '2']
            self.assertEqual(cell.value, header)
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.alignment.horizontal, 'center')
            self.assertEqual(cell.alignment.vertical, 'center')

    def test_dateProcess(self):
        date_str = "12-12-2022"
        processed_date = dateProcess(date_str)
        self.assertEqual(processed_date, "12-12-2022")

        date_str = "12/Dec/2022"
        processed_date = dateProcess(date_str)
        self.assertEqual(processed_date, "12/Dec/2022")

        date_str = "Invalid Date"
        processed_date = dateProcess(date_str)
        self.assertEqual(processed_date, "Invalid Date")

    

    def test_jsonProcess(self):
        # Mock a JSON file
        test_json_data = {"key": "value"}
        with patch('builtins.open', mock_open(read_data=json.dumps(test_json_data))):
            with patch('os.path.isfile', return_value=True):
                result = jsonProcess("fake_path.json")
                self.assertEqual(result, test_json_data)

        # Test non-existent file
        with patch('os.path.isfile', return_value=False):
            result = jsonProcess("non_existent_file.json")
            self.assertEqual(result, "non_existent_file.json")

    @patch('excel_management.load_workbook')
    def test_manuProcess(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Manufacture'
        ws.append(['Keyword', 'Value'])
        ws.append(['14JA1-4', 'Manufacturer 1'])
        ws.append(['XYZ123', 'Manufacturer 2'])
        
        # Mock load_workbook
        mock_load_workbook.return_value = wb
        
        self.assertEqual(manuProcess('14JA1-4'), 'Manufacturer 1')
        self.assertEqual(manuProcess('XYZ123'), 'Manufacturer 2')
        self.assertEqual(manuProcess('ABC'), None)
    
    def test_dateProcess(self):
        self.assertEqual(dateProcess('12-03-2020'), '12-03-2020')
        self.assertEqual(dateProcess('12-Mar-2020'), '12-Mar-2020')
        self.assertEqual(dateProcess('January 5, 2021'), 'January 5, 2021')
        self.assertEqual(dateProcess('Random Text'), 'Random Text')
        self.assertIsNot(dateProcess('2020/03/12'), '03/12/2020')

    def test_swlProcess(self):
        wb = Workbook()
        ws = wb.active
        
        error_pages = []
        swlProcess(ws, ['12.5kg Note1', '15kg Note2'], 1, error_pages)
        
        self.assertEqual(ws.cell(row=1, column=6).value, '12.5')
        self.assertEqual(ws.cell(row=1, column=7).value, 'kg')
        self.assertEqual(ws.cell(row=1, column=8).value, 'Note1')
        
        self.assertEqual(ws.cell(row=2, column=6).value, '15')
        self.assertEqual(ws.cell(row=2, column=7).value, 'kg')
        self.assertEqual(ws.cell(row=2, column=8).value, 'Note2')
    
    @patch('excel_management.load_workbook')
    def test_modelProcess(self, mock_load_workbook):
        wb = Workbook()
        ws = wb.create_sheet(title='Model')
        ws.append(['Keyword', 'Value'])
        ws.append(['Model1', 'Manufacturer1'])
        ws.append(['Model2', 'Manufacturer2'])
        
        mock_load_workbook.return_value = wb
        
        self.assertEqual(modelProcess('Model1'), ('Model1', 'Manufacturer1'))
        self.assertEqual(modelProcess('Model2'), ('Model2', 'Manufacturer2'))
        self.assertIsNone(modelProcess('UnknownModel'))

    @patch('excel_management.modelProcess')
    @patch('excel_management.dateProcess')
    def test_insertData(self, mock_dateProcess, mock_modelProcess):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Extraction Data'
        
        mock_dateProcess.side_effect = lambda x: x
        mock_modelProcess.side_effect = lambda x: (x, 'MockManufacturer') if x == 'Model1' else None
        
        extracted_data = {
            'SWL': ['10kg', '20kg Note'],
            'Previous Inspection': ['12-03-2020', '14-04-2021'],
            "Next Inspection Due Date" : ['12-03-2021', '12-03-2022'],
            'Model': ['Model1', 'UnknownModel']
        }
        
        error_pages = []
        
        insertData(extracted_data, ws, error_pages)
        
        self.assertEqual(ws.cell(row=3, column=ord("E") - 64).value, 'Model1')
        self.assertEqual(ws.cell(row=3, column=ord("I") - 64).value, 'MockManufacturer')
        self.assertEqual(ws.cell(row=4, column=ord("E") - 64).value, 'UnknownModel')
        self.assertEqual(ws.cell(row=3, column=ord("M") - 64).value, '12-03-2020')
        self.assertEqual(ws.cell(row=4, column=ord("M") - 64).value, '14-04-2021')
        self.assertEqual(ws.cell(row=3, column=ord("F") - 64).value, '10')
        self.assertEqual(ws.cell(row=3, column=ord("G") - 64).value, 'kg')
        self.assertEqual(ws.cell(row=4, column=ord("F") - 64).value, '20')
        self.assertEqual(ws.cell(row=4, column=ord("G") - 64).value, 'kg')
        self.assertEqual(ws.cell(row=4, column=ord("H") - 64).value, 'Note')

    def test_insertError(self):
        wb = Workbook()
        error_pages = ["Error 1", "Error 2"]
        
        insertError(wb, error_pages)
        
        ws = wb["Errors"]
        
        self.assertEqual(ws.cell(row=2, column=ord("A") - 64).value, 1)
        self.assertEqual(ws.cell(row=2, column=ord("B") - 64).value, "Error 1")
        self.assertEqual(ws.cell(row=3, column=ord("A") - 64).value, 2)
        self.assertEqual(ws.cell(row=3, column=ord("B") - 64).value, "Error 2")

    @patch('excel_management.openpyxl.Workbook')
    @patch('excel_management.insertData')
    @patch('excel_management.insertError')
    def test_create_excel(self, mock_insertError, mock_insertData, mock_Workbook):
        extracted_data = {
            'SWL': ['10kg', '20kg Note'],
            'Date': ['12-03-2020', '14-04-2021'],
            'Model': ['Model1', 'UnknownModel']
        }
        filename = 'test.xlsx'
        client = 'Client1'
        page_errors = {}

        wb = MagicMock()
        ws = MagicMock()
        wb.active = ws
        mock_Workbook.return_value = wb

        create_excel(extracted_data, filename, client, page_errors)

        
        mock_insertData.assert_called_once_with(extracted_data, ws, [])
        mock_insertError.assert_called_once_with(wb, [])
        wb.save.assert_called_once_with(filename)
        wb.close.assert_called_once()
    

if __name__ == '__main__':
    unittest.main()