import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json
import re

column_mapping = {
    "Id Number": "A",
    "RFID": "B",
    "Item Category": "C",
    "Item Description": "D",
    "Model": "E",
    "SWL Value": "F",
    "SWL Unit": "G",
    "SWL Note": "H",
    "Manufacturer": "I",
    "Certificate No": "J",
    "Location": "K",
    "Detailed Location ": "L",
    "Previous Inspection": "M",
    "Next Inspection Due Date": "N",
    "Fit For Purpose Y/N": "O",
    "Status" : "P",
    "Provider Identification": "Q",
    "Errors": "R"
}

def preProcess(sheet_data, column_mapping):
    for header, column in column_mapping.items():
            cell = sheet_data[column + '2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the text
            cell.border = Border(bottom=Side(border_style='thin'))  # Add a thin border at the bottom

            # Adjust column width to fit the header text
            column_width = max(len(header), 10)  # Set a minimum width of 10 characters
            sheet_data.column_dimensions[column].width = column_width

def insertData(extracted_data, sheet_data):
    nameList = list(extracted_data.keys())  
    row_idx = 3
    for name in nameList:
        if name == "SWL":
                swlVlaue = extracted_data.get(name)
                swlProcess(sheet_data, swlVlaue)
        
        else:
            column_name = column_mapping.get(name)  
            if column_name:
                row_idx = 3
                
                for val in extracted_data.get(name, []):
                    if isinstance(val, str):  
                        if column_name == "M" or column_name == "N":
                            val = dateProcess(val)
                        if column_name == "E":
                            modelData = extracted_data.get(name)
                            modelTuple = modelProcess(modelData)
                            val = modelTuple[0]
                            manu = modelTuple[1] 
                            sheet_data.cell(row=row_idx, column=ord("I") - 64, value=manu)
                        if column_name == "I":
                            manuData = extracted_data.get(name)
                            val = modelProcess(manuData)
                        sheet_data.cell(row=row_idx, column=ord(column_name) - 64, value=val)
                        row_idx += 1 


def insertError(workbook, page_errors):
    # Create a sheet for errors
        sheet_errors = workbook.create_sheet(title="Errors")  # Create a new worksheet
        sheet_errors.append(["No", "Error"])  # Write column headers

        # Write errors to the worksheet
        idx = 2
        for value in page_errors:
            if isinstance(value, str):
                sheet_errors.cell(row=idx, column=ord("A") - 64, value=idx - 1)
                sheet_errors.cell(row=idx, column=ord("B") - 64, value=value)
            idx += 1
            

def create_excel(extracted_data: dict, filename: str, client: str, page_errors: dict):
        
    try:
        print("<--------------Creating new excel------------------>")
        workbook = openpyxl.Workbook()  # Create a new Workbook

        # Create a sheet for extracted data
        sheet_data = workbook.active
        sheet_data.title = "Extraction Data"  # Set sheet name

        sheet_data['A1'] = "Rig-Ware import v2"
        sheet_data['B1'] = client
        sheet_data['C1'] = "CreateLocations=No"

        # Write column headers for extracted data sheet
        preProcess(sheet_data, column_mapping)
        
        # Insert data from json 
        insertData(extracted_data, sheet_data)

        # Insert error infomration
        insertError(workbook, page_errors)

        workbook.save(filename)  # Save the workbook with the provided filename
        workbook.close()
        print("<-------------- Excel created successfully ------------------>")
    except Exception as e:
       print(f"An error occurred in excel creation: {e}")

def modelProcess(modelData):
    workbook = load_workbook("database/Full_list_of_Manufacturers_and_Models.xlsx")
    model_sheet = workbook['Model']
    for row in model_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
        if keyword in modelData:
            val = keyword
            
    return val, value

def manuProcess(manuData):
    workbook = load_workbook("database/Full_list_of_Manufacturers_and_Models.xlsx")
    manufacturer_sheet = workbook['Manufacture']
    for row in manufacturer_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        value = row[1]
    if keyword in manuData:
            val = value
    return val



def dateProcess(val):
    pattern = r'\b\d{2}[-/](?:\d{2}|[A-Za-z]{3})[-/]\d{4}\b|[A-Za-z]+(?:\s+[A-Za-z]+)*'
    val = re.findall(pattern, val)
    if len(val) != 0:
        val = val[0]
    return val


def swlProcess(sheet_data, swlValue):
    
    valPattern = r'\d+(?:\.\d+)?'
    unitPattern = r'[a-zA-Z]+'
    notePattern = r'[\s\n]+(\S.*)'
    
    idx = 3
    for val in swlValue:
        if isinstance(val, str):
            swlVal = re.findall(valPattern, val)[0]
            sheet_data.cell(row=idx, column=ord("F") - 64, value=swlVal)
            
            swlUnit = re.findall(unitPattern, val)[0]
           
            sheet_data.cell(row=idx, column=ord("G") - 64, value=swlUnit)
            
            swlNote = re.findall(notePattern, val)
           
            if(len(swlNote) > 0) and (swlNote[0] != swlUnit):
                swlNote = swlNote[0]
                sheet_data.cell(row=idx, column=ord("H") - 64, value=swlNote)
            idx += 1

        


def jsonProcess():
    with open('coords.json', 'r') as file:
        data = json.load(file)
    
    return data


if __name__ == "__main__":
    data = jsonProcess()
    error_page = data.get("Error Page")
    create_excel(data, "output.xlsx", "Client Name", error_page)
